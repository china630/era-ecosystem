# -*- coding: utf-8 -*-
import glob, os, io, json, collections, datetime, re
import openpyxl

REPO = r"d:\My Projects\era-ecosystem"
OUT = os.path.join(REPO, "era-clinic", "prisma", "seed-data", "nafta")
WO = r"C:\Users\ASUS G752VT\Downloads\WO"

# ---- transliteration for stable ascii codes ----
TR = {"ə":"e","Ə":"E","ç":"c","Ç":"C","ş":"s","Ş":"S","ğ":"g","Ğ":"G","ı":"i","İ":"I","ö":"o","Ö":"O","ü":"u","Ü":"U"}
def slug(s):
    s = "".join(TR.get(ch, ch) for ch in s)
    s = s.upper()
    s = re.sub(r"[^A-Z0-9]+", "-", s).strip("-")
    return s[:40]

# ---- full price catalog from prices xlsx ----
PRICES = openpyxl.load_workbook(r"C:\Users\ASUS G752VT\Downloads\nafta_sanatorium_prices.xlsx", read_only=True, data_only=True)
catalog = []
for ws in PRICES.worksheets:
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or r[0] is None:
            continue
        az = str(r[0]).strip()
        ru = str(r[1]).strip() if len(r) > 1 and r[1] is not None else az
        price = None
        if len(r) > 2 and r[2] is not None:
            m = re.search(r"\d+", str(r[2]))
            price = int(m.group()) if m else None
        catalog.append({"code": "SVC-" + slug(az), "nameAz": az, "nameRu": ru,
                         "price": price, "sheet": ws.title})
with io.open(os.path.join(OUT, "catalog.json"), "w", encoding="utf-8") as f:
    json.dump(catalog, f, ensure_ascii=False, indent=1)

# ---- Randevular procedure -> (nameAz canonical, price, duration) ----
# price from physio/vanna/policlinika sheets
M = {
 "Amplipuls": ("Amplipuls", 11, 20),
 "Ultrafonoforez (Naftalan yağıyla)": ("Ultrafonoforez", 12, 20),
 "Ultrafonoforez (Gellə)": ("Ultrafonoforez", 12, 20),
 "Ozonterapiya": ("Ozonterapiya", 17, 20),
 "Darsonval": ("Darsonval", 11, 20),
 "Lazerterapiya": ("Lazerterapiya", 12, 20),
 "Naftalan vannası (Qadın)": ("Tam bədən naftalan vannası", 14, 20),
 "Naftalan vannası (Kişi)": ("Tam bədən naftalan vannası", 14, 20),
 "Solyuks": ("Sollyuks", 14, 20),
 "Parafin Yuxarı nahiyə": ("Parafinoterapiya (yuxarı ətraf)", 11, 20),
 "Maqnitoterapiya": ("Maqnitoterapiya", 12, 20),
 "İnfraqırmızı": ("İnfraqırmızı", 14, 20),
 "4 kamera vanna": ("4 kameralı naftalan vannası", 12, 20),
 "Massaj 30": ("Klasik massaj (30 dəqiqə)", 19, 30),
 "Elektroforez": ("Elektroterapiya", 14, 20),
 "UFB terapiya": ("UFB terapiya", 20, 20),
 "Parafin Aşağı nahiyə": ("Parafinoterapiya (aşağı ətraf)", 12, 20),
 "Qısa dalğa terapiya UVÇ": ("Qısa dalğa terapiya UVÇ", 12, 20),
 "Parafin Kürək - onurğa": ("Parafinoterapiya (boyun kürək)", 14, 20),
 "Super induktiv terapiya": ("Super Inductive system terapiyası", 30, 20),
 "Vakumterapiya": ("Vakuumterapiya", 11, 20),
 "Bükmə": ("Bükmə", 20, 20),
 "Limfodrenaj": ("Limfodrenaj", 16, 30),
 "4 kamera hidroqalvanizasiya": ("4 kameralı hidroqalvanizasiya", 15, 20),
 "Massaj 15": ("Klasik massaj (15 dəqiqə)", 16, 15),
 "Hidromasaj vanna": ("Hidromasaj vannası", 15, 20),
 "Trunda burun": ("Turunda (burun və qulaq)", 12, 15),
 "İnqalyasiya": ("İnqalyasiya", 10, 15),
 "Zərbə dalğa": ("Zərbə dalğa terapiya", 24, 20),
 "Yod-brom vanna": ("Yod brom vannası", 18, 20),
 "Hidrokоlon ( bitki çayı ilə)": ("Hidrokolonoterapiya (Bitki çayı ilə)", 35, 40),
 "Hidrokolon": ("Hidrokolonoterapiya", 30, 40),
 "Karboksiterapiya": ("Karboksiterapiya", 18, 20),
 "Parafin bütün bədən": ("Parafinoterapiya (bütün bədən)", 18, 30),
 "Limfa düyünləri": ("Limfodrenaj", 16, 20),
 "Massaj 30 (test)": ("Klasik massaj (30 dəqiqə)", 19, 30),
 "Mikroklizma": ("Uroloji mikroklizma", 14, 15),
 "Uroloji vibro lazer": ("Uroloji Vibro lazer", 12, 20),
 "Fitoterapiya ( boçka )": ("Fito terapiya (boçka)", 16, 20),
}

daily = sorted(glob.glob(os.path.join(WO, "Randevular_2026-07-1*.xlsx")))
rows = []  # (date, seq, patient, proc_raw, cabinet)
patients = set(); cabinets = set(); used_proc = {}
for f in daily:
    d = os.path.basename(f).replace("Randevular_", "").replace(".xlsx", "")
    dt = datetime.date.fromisoformat(d)
    if dt.weekday() == 6:  # Sunday: clinic closed
        continue
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or r[1] is None:
            continue
        patient = str(r[1]).strip()
        proc = str(r[2]).strip()
        cab = str(r[3]).strip()
        seq = r[0] if isinstance(r[0], int) else i
        if "test" in proc.lower():
            continue  # drop test rows
        if proc not in M:
            continue
        patients.add(patient); cabinets.add(cab)
        canon, price, dur = M[proc]
        code = "SVC-" + slug(canon)
        used_proc[code] = {"code": code, "name": canon, "price": price, "durationMin": dur}
        rows.append((d, seq, patient, proc, canon, code, price, dur, cab))
    wb.close()

# patients fixture (sorted, stable refCodes)
pat_sorted = sorted(patients, key=lambda s: s.lower())
pat_code = {}
plist = []
for idx, name in enumerate(pat_sorted, start=1):
    rc = "NAFTA-P%04d" % idx
    pat_code[name] = rc
    plist.append({"refCode": rc, "fullName": name})
with io.open(os.path.join(OUT, "patients.json"), "w", encoding="utf-8") as f:
    json.dump(plist, f, ensure_ascii=False, indent=1)

# cabinets fixture
cab_sorted = sorted(cabinets, key=lambda s: s.lower())
cab_code = {}
clist = []
for name in cab_sorted:
    cc = "CAB-" + slug(name)
    cab_code[name] = cc
    clist.append({"code": cc, "name": name})
with io.open(os.path.join(OUT, "cabinets.json"), "w", encoding="utf-8") as f:
    json.dump(clist, f, ensure_ascii=False, indent=1)

# procedure types fixture
with io.open(os.path.join(OUT, "procedure-types.json"), "w", encoding="utf-8") as f:
    json.dump(sorted(used_proc.values(), key=lambda x: x["code"]), f, ensure_ascii=False, indent=1)

# assign times per (date, cabinet) sequentially, ignoring source times
TODAY = datetime.date(2026, 7, 18)
NOWH = 11
by_dc = collections.defaultdict(list)
for row in rows:
    by_dc[(row[0], row[8])].append(row)

appts = []
for (d, cab), lst in by_dc.items():
    lst.sort(key=lambda x: x[1] if isinstance(x[1], int) else 0)
    dt = datetime.date.fromisoformat(d)
    start_h = 8 if dt.weekday() == 5 else 9  # Saturday earlier
    cur = datetime.datetime(dt.year, dt.month, dt.day, start_h, 0)
    for row in lst:
        dur = row[7]
        # skip lunch 13:00-14:00
        if cur.hour == 13:
            cur = cur.replace(hour=14, minute=0)
        t = cur.strftime("%H:%M")
        if dt < TODAY:
            status = "COMPLETED"
        else:
            status = "CHECKED_IN" if cur.hour < NOWH else "SCHEDULED"
        appts.append({
            "date": d, "time": t,
            "patientRefCode": pat_code[row[2]],
            "procedureCode": row[5], "procedureName": row[4],
            "cabinetCode": cab_code[row[8]], "amount": row[6], "status": status,
        })
        cur += datetime.timedelta(minutes=dur)
appts.sort(key=lambda a: (a["date"], a["cabinetCode"], a["time"]))
with io.open(os.path.join(OUT, "appointments.json"), "w", encoding="utf-8") as f:
    json.dump(appts, f, ensure_ascii=False, indent=1)

print("catalog", len(catalog), "procTypes", len(used_proc), "patients", len(plist),
      "cabinets", len(clist), "appointments", len(appts))