#!/usr/bin/env python3
"""
Собирает структуру Excel-экспортов из doc/Elektraweb в сводный манифест.

Для каждого .xlsx читает заголовки (первую строку), имя листа и число строк данных.
В манифесте обязательно указывается папка модуля, из которой взят файл.

Зависимости:
  pip install openpyxl

Запуск:
  python scripts/build-elektraweb-manifest.py
  python scripts/build-elektraweb-manifest.py --module "14 POS"
  python scripts/build-elektraweb-manifest.py --skip-row-count
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

try:
    import openpyxl
except ImportError:
    print(
        "Missing dependencies. Install with:\n"
        "  pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
ELEKTRAWEB_DIR = ROOT / "doc" / "Elektraweb"
CSV_OUT = ELEKTRAWEB_DIR / "elektraweb-screens-manifest.csv"
JSON_OUT = ELEKTRAWEB_DIR / "elektraweb-screens-manifest.json"

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"m": MAIN_NS}

FILENAME_RE = re.compile(
    r"^(.+?)\.(\d{4}-\d{2}-\d{2})\.(\d{2}-\d{2}-\d{2})\.(.+)\.xlsx$",
    re.IGNORECASE,
)


def folder_to_label(folder_name: str) -> str:
    match = re.match(r"^\d+\s+(.+)$", folder_name.strip())
    return match.group(1) if match else folder_name.strip()


def natural_sort_key(text: str) -> list:
    parts = re.split(r"(\d+)", text.lower())
    return [int(part) if part.isdigit() else part for part in parts]


def parse_filename(path: Path) -> dict[str, str | None]:
    match = FILENAME_RE.match(path.name)
    if not match:
        return {
            "screen_name": path.stem,
            "exported_date": None,
            "exported_time": None,
            "hotel_name": None,
        }

    return {
        "screen_name": match.group(1).strip(),
        "exported_date": match.group(2),
        "exported_time": match.group(3).replace("-", ":"),
        "hotel_name": match.group(4).strip(),
    }


def normalize_header(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_headers(row: tuple[Any, ...] | list[Any]) -> list[str]:
    headers: list[str] = []
    for cell in row:
        header = normalize_header(cell)
        if header is not None:
            headers.append(header)
    return headers


def row_has_data(row: tuple[Any, ...]) -> bool:
    return any(cell is not None and str(cell).strip() != "" for cell in row)


def read_workbook_openpyxl(path: Path, skip_row_count: bool) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        sheet_names = list(workbook.sheetnames)
        header_row = next(sheet.iter_rows(max_row=1, values_only=True), ())
        headers = normalize_headers(header_row)

        data_row_count = 0
        if not skip_row_count:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row_has_data(row):
                    data_row_count += 1

        return {
            "sheet_name": sheet.title,
            "sheet_names": sheet_names,
            "columns": headers,
            "data_row_count": data_row_count,
            "read_method": "openpyxl",
            "read_error": None,
        }
    finally:
        workbook.close()


def read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []

    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings: list[str] = []
    for item in root.findall("m:si", NS):
        text_parts = [node.text or "" for node in item.findall(".//m:t", NS)]
        strings.append("".join(text_parts))
    return strings


def resolve_first_sheet_path(archive: zipfile.ZipFile) -> str:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    first_sheet = workbook.find("m:sheets/m:sheet", NS)
    if first_sheet is None:
        raise ValueError("Workbook has no sheets")

    rel_id = first_sheet.attrib.get(f"{{{REL_NS}}}id")
    if not rel_id:
        raise ValueError("First sheet has no relationship id")

    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    for rel in relationships.findall(f"{{{PKG_REL_NS}}}Relationship"):
        if rel.attrib.get("Id") == rel_id:
            target = rel.attrib["Target"]
            if target.startswith("/"):
                return target.lstrip("/")
            return f"xl/{target.lstrip('/')}"

    raise ValueError(f"Sheet relationship not found: {rel_id}")


def cell_value(cell: ET.Element, shared_strings: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    value_node = cell.find("m:v", NS)

    if cell_type == "inlineStr":
        text_node = cell.find(".//m:t", NS)
        return text_node.text if text_node is not None else None

    if value_node is None or value_node.text is None:
        return None

    raw = value_node.text
    if cell_type == "s":
        index = int(raw)
        return shared_strings[index] if 0 <= index < len(shared_strings) else raw
    if cell_type == "b":
        return raw == "1"
    return raw


def read_workbook_xml(path: Path, skip_row_count: bool) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = read_shared_strings(archive)
        sheet_path = resolve_first_sheet_path(archive)
        sheet_root = ET.fromstring(archive.read(sheet_path))

        sheet_data = sheet_root.find("m:sheetData", NS)
        if sheet_data is None:
            raise ValueError("Sheet data is missing")

        rows = sheet_data.findall("m:row", NS)
        header_cells = rows[0].findall("m:c", NS) if rows else []
        headers = normalize_headers([cell_value(cell, shared_strings) for cell in header_cells])

        data_row_count = 0
        if not skip_row_count:
            for row in rows[1:]:
                values = [cell_value(cell, shared_strings) for cell in row.findall("m:c", NS)]
                if row_has_data(tuple(values)):
                    data_row_count += 1

        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        first_sheet = workbook.find("m:sheets/m:sheet", NS)
        sheet_name = first_sheet.attrib.get("name", "") if first_sheet is not None else ""
        sheet_names = [
            sheet.attrib.get("name", "")
            for sheet in workbook.findall("m:sheets/m:sheet", NS)
        ]

        return {
            "sheet_name": sheet_name,
            "sheet_names": sheet_names,
            "columns": headers,
            "data_row_count": data_row_count,
            "read_method": "xml-fallback",
            "read_error": None,
        }


def read_workbook(path: Path, skip_row_count: bool) -> dict[str, Any]:
    try:
        return read_workbook_openpyxl(path, skip_row_count)
    except Exception as exc:
        try:
            result = read_workbook_xml(path, skip_row_count)
            result["read_error"] = f"openpyxl: {exc}"
            return result
        except Exception as fallback_exc:
            return {
                "sheet_name": None,
                "sheet_names": [],
                "columns": [],
                "data_row_count": None,
                "read_method": "failed",
                "read_error": f"openpyxl: {exc}; xml-fallback: {fallback_exc}",
            }


def collect_excel_files(folder: Path) -> list[Path]:
    files = [path for path in folder.iterdir() if path.is_file() and path.suffix.lower() == ".xlsx"]
    return sorted(files, key=lambda path: natural_sort_key(path.name))


def build_record(module_folder: str, path: Path, skip_row_count: bool) -> dict[str, Any]:
    meta = parse_filename(path)
    workbook = read_workbook(path, skip_row_count=skip_row_count)
    columns = workbook["columns"]
    data_row_count = workbook["data_row_count"]

    return {
        "module_folder": module_folder,
        "module_label": folder_to_label(module_folder),
        "source_file": path.name,
        "source_path": str(path.relative_to(ROOT)).replace("\\", "/"),
        "screen_name": meta["screen_name"],
        "exported_date": meta["exported_date"],
        "exported_time": meta["exported_time"],
        "hotel_name": meta["hotel_name"],
        "sheet_name": workbook["sheet_name"],
        "sheet_names": workbook["sheet_names"],
        "column_count": len(columns),
        "data_row_count": data_row_count,
        "is_empty": True if data_row_count == 0 else False if data_row_count else None,
        "columns": columns,
        "read_method": workbook["read_method"],
        "read_error": workbook["read_error"],
    }


def write_csv(records: list[dict[str, Any]], output_path: Path) -> None:
    fieldnames = [
        "module_folder",
        "module_label",
        "source_file",
        "source_path",
        "screen_name",
        "exported_date",
        "exported_time",
        "hotel_name",
        "sheet_name",
        "column_count",
        "data_row_count",
        "is_empty",
        "columns_json",
        "read_method",
        "read_error",
    ]

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter=";", lineterminator="\n")
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "module_folder": record["module_folder"],
                    "module_label": record["module_label"],
                    "source_file": record["source_file"],
                    "source_path": record["source_path"],
                    "screen_name": record["screen_name"],
                    "exported_date": record["exported_date"] or "",
                    "exported_time": record["exported_time"] or "",
                    "hotel_name": record["hotel_name"] or "",
                    "sheet_name": record["sheet_name"] or "",
                    "column_count": record["column_count"],
                    "data_row_count": "" if record["data_row_count"] is None else record["data_row_count"],
                    "is_empty": "" if record["is_empty"] is None else str(record["is_empty"]).lower(),
                    "columns_json": json.dumps(record["columns"], ensure_ascii=False),
                    "read_method": record["read_method"],
                    "read_error": record["read_error"] or "",
                }
            )


def write_json(records: list[dict[str, Any]], output_path: Path) -> None:
    module_map: dict[str, list[dict[str, Any]]] = {}
    for record in records:
        module_map.setdefault(record["module_folder"], []).append(record)

    modules = []
    for module_folder in sorted(module_map, key=natural_sort_key):
        files = module_map[module_folder]
        modules.append(
            {
                "module_folder": module_folder,
                "module_label": folder_to_label(module_folder),
                "file_count": len(files),
                "files": files,
            }
        )

    payload = {
        "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "source_root": str(ELEKTRAWEB_DIR.relative_to(ROOT)).replace("\\", "/"),
        "summary": {
            "module_count": len(modules),
            "file_count": len(records),
            "empty_files": sum(1 for record in records if record["is_empty"] is True),
            "files_with_data": sum(
                1 for record in records if isinstance(record["data_row_count"], int) and record["data_row_count"] > 0
            ),
            "failed_files": sum(1 for record in records if record["read_method"] == "failed"),
            "by_module": {
                module["module_folder"]: module["file_count"] for module in modules
            },
        },
        "records": records,
        "modules": modules,
    }

    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build consolidated CSV/JSON manifest from Elektraweb Excel exports.",
    )
    parser.add_argument(
        "--module",
        help='Process only one subfolder, e.g. "14 POS" or "17 SPA".',
    )
    parser.add_argument(
        "--skip-row-count",
        action="store_true",
        help="Read only headers and sheet names (faster).",
    )
    parser.add_argument(
        "--csv-out",
        type=Path,
        default=CSV_OUT,
        help=f"CSV output path (default: {CSV_OUT.name}).",
    )
    parser.add_argument(
        "--json-out",
        type=Path,
        default=JSON_OUT,
        help=f"JSON output path (default: {JSON_OUT.name}).",
    )
    args = parser.parse_args()

    if not ELEKTRAWEB_DIR.is_dir():
        print(f"Folder not found: {ELEKTRAWEB_DIR}", file=sys.stderr)
        return 1

    if args.module:
        folders = [ELEKTRAWEB_DIR / args.module]
        if not folders[0].is_dir():
            print(f"Module folder not found: {folders[0]}", file=sys.stderr)
            return 1
    else:
        folders = sorted(
            [path for path in ELEKTRAWEB_DIR.iterdir() if path.is_dir()],
            key=lambda path: natural_sort_key(path.name),
        )

    records: list[dict[str, Any]] = []
    print(f"Source: {ELEKTRAWEB_DIR}")
    print()

    for folder in folders:
        excel_files = collect_excel_files(folder)
        print(f"  {folder.name}: {len(excel_files)} xlsx")
        for excel_path in excel_files:
            records.append(build_record(folder.name, excel_path, skip_row_count=args.skip_row_count))

    args.csv_out.parent.mkdir(parents=True, exist_ok=True)
    args.json_out.parent.mkdir(parents=True, exist_ok=True)
    write_csv(records, args.csv_out)
    write_json(records, args.json_out)

    failed = [record for record in records if record["read_method"] == "failed"]
    fallback = [record for record in records if record["read_method"] == "xml-fallback"]

    print()
    print(f"CSV : {args.csv_out}")
    print(f"JSON: {args.json_out}")
    print(f"Files processed: {len(records)}")
    print(f"Modules: {len(folders)}")
    if fallback:
        print(f"XML fallback used: {len(fallback)}")
    if failed:
        print(f"Failed: {len(failed)}", file=sys.stderr)
        for record in failed:
            print(f"  - [{record['module_folder']}] {record['source_file']}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
